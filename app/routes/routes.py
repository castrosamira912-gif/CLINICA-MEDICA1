import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    send_file
)

from flask_login import (
    login_user,
    logout_user,
    login_required,
    current_user
)

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from app import db

from app.models.medico import Medico
from app.models.paciente import Paciente
from app.models.consulta import Consulta
from app.models.usuario import Usuario
from app.models.cita import Cita


main = Blueprint('main', __name__)


# =========================
# LOGIN
# =========================

@main.route('/registro', methods=['GET', 'POST'])
def registro():

    if request.method == 'POST':

        username = request.form['username']

        password = generate_password_hash(
            request.form['password']
        )

        nuevo_usuario = Usuario(
            username=username,
            password=password
        )

        db.session.add(nuevo_usuario)

        db.session.commit()

        return redirect(url_for('main.login'))

    return render_template('registro.html')


@main.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        username = request.form['username']

        password = request.form['password']

        usuario = Usuario.query.filter_by(
            username=username
        ).first()

        if usuario and check_password_hash(
            usuario.password,
            password
        ):

            login_user(usuario)

            return redirect(url_for('main.index'))

    return render_template('login.html')


@main.route('/logout')
@login_required
def logout():

    logout_user()

    return redirect(url_for('main.login'))


# =========================
# INICIO
# =========================

@main.route('/')
@login_required
def index():
    total_medicos = Medico.query.count()
    total_pacientes = Paciente.query.count()
    total_consultas = Consulta.query.count()

    return render_template(
        'index.html',
        total_medicos=total_medicos,
        total_pacientes=total_pacientes,
        total_consultas=total_consultas
    )


# =========================
# MEDICOS
# =========================

@main.route('/medicos')
@login_required
def medicos():

    lista_medicos = Medico.query.all()

    return render_template(
        'medicos.html',
        medicos=lista_medicos
    )


@main.route('/crear_medico', methods=['GET', 'POST'])
@login_required
def crear_medico():

    if request.method == 'POST':

        nombre = request.form['nombre']
        especialidad = request.form['especialidad']
        telefono = request.form['telefono']
        correo = request.form['correo']

        nuevo_medico = Medico(
            nombre=nombre,
            especialidad=especialidad,
            telefono=telefono,
            correo=correo
        )

        db.session.add(nuevo_medico)

        db.session.commit()

        return redirect(url_for('main.medicos'))

    return render_template('crear_medico.html')


@main.route('/editar_medico/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_medico(id):

    medico = Medico.query.get_or_404(id)

    if request.method == 'POST':

        medico.nombre = request.form['nombre']
        medico.especialidad = request.form['especialidad']
        medico.telefono = request.form['telefono']
        medico.correo = request.form['correo']

        db.session.commit()

        return redirect(url_for('main.medicos'))

    return render_template(
        'editar_medico.html',
        medico=medico
    )


@main.route('/eliminar_medico/<int:id>')
@login_required
def eliminar_medico(id):

    medico = Medico.query.get_or_404(id)

    if medico.consultas:

        return '''
        <h2>
        No puedes eliminar este médico
        porque tiene consultas registradas.
        </h2>

        <a href="/medicos">
        Volver
        </a>
        '''

    db.session.delete(medico)

    db.session.commit()

    return redirect(url_for('main.medicos'))


# =========================
# PACIENTES
# =========================

@main.route('/pacientes')
@login_required
def pacientes():

    lista_pacientes = Paciente.query.all()

    return render_template(
        'pacientes.html',
        pacientes=lista_pacientes
    )


@main.route('/crear_paciente', methods=['GET', 'POST'])
@login_required
def crear_paciente():

    if request.method == 'POST':

        nombre = request.form['nombre']
        edad = request.form['edad']
        direccion = request.form['direccion']
        telefono = request.form['telefono']

        nuevo_paciente = Paciente(
            nombre=nombre,
            edad=edad,
            direccion=direccion,
            telefono=telefono
        )

        db.session.add(nuevo_paciente)

        db.session.commit()

        return redirect(url_for('main.pacientes'))

    return render_template('crear_paciente.html')


# =========================
# CONSULTAS
# =========================

@main.route('/consultas')
@login_required
def consultas():

    fecha = request.args.get('fecha')

    if fecha:

        lista_consultas = Consulta.query.filter_by(
            fecha=fecha
        ).all()

    else:

        lista_consultas = Consulta.query.all()

    return render_template(
        'consultas.html',
        consultas=lista_consultas
    )


@main.route('/crear_consulta', methods=['GET', 'POST'])
@login_required
def crear_consulta():

    medicos = Medico.query.all()

    pacientes = Paciente.query.all()

    if request.method == 'POST':

        fecha = request.form['fecha']

        diagnostico = request.form['diagnostico']

        tratamiento = request.form['tratamiento']

        id_medico = request.form['id_medico']

        id_paciente = request.form['id_paciente']

        nueva_consulta = Consulta(
            fecha=fecha,
            diagnostico=diagnostico,
            tratamiento=tratamiento,
            id_medico=id_medico,
            id_paciente=id_paciente
        )

        db.session.add(nueva_consulta)

        db.session.commit()

        return redirect(url_for('main.consultas'))

    return render_template(
        'crear_consulta.html',
        medicos=medicos,
        pacientes=pacientes
    )


# =========================
# CITAS
# =========================

@main.route('/citas')
@login_required
def citas():

    lista_citas = Cita.query.all()

    return render_template(
        'citas.html',
        citas=lista_citas
    )


@main.route('/crear_cita', methods=['GET', 'POST'])
@login_required
def crear_cita():

    medicos = Medico.query.all()

    pacientes = Paciente.query.all()

    if request.method == 'POST':

        nueva_cita = Cita(

            fecha=request.form['fecha'],

            hora=request.form['hora'],

            motivo=request.form['motivo'],

            id_medico=request.form['id_medico'],

            id_paciente=request.form['id_paciente']
        )

        db.session.add(nueva_cita)

        db.session.commit()

        return redirect(url_for('main.citas'))

    return render_template(
        'crear_cita.html',
        medicos=medicos,
        pacientes=pacientes
    )


# =========================
# HISTORIAL MEDICO
# =========================

@main.route('/historial/<int:id>')
@login_required
def historial(id):

    paciente = Paciente.query.get_or_404(id)

    consultas = Consulta.query.filter_by(
        id_paciente=id
    ).all()

    return render_template(
        'historial.html',
        paciente=paciente,
        consultas=consultas
    )


# =========================
# EXPORTAR EXCEL
# =========================

@main.route('/exportar_consultas')
@login_required
def exportar_consultas():

    consultas = Consulta.query.all()

    datos = []

    for c in consultas:

        datos.append({

            'Fecha': c.fecha,

            'Paciente': c.paciente.nombre,

            'Médico': c.medico.nombre,

            'Diagnóstico': c.diagnostico,

            'Tratamiento': c.tratamiento
        })

    df = pd.DataFrame(datos)

    ruta_archivo = os.path.join(
        os.getcwd(),
        'Registro_Clinico.xlsx'
    )

    df.to_excel(
        ruta_archivo,
        index=False
    )

    wb = load_workbook(ruta_archivo)

    ws = wb.active

    color_encabezado = PatternFill(
        start_color='0D6EFD',
        end_color='0D6EFD',
        fill_type='solid'
    )

    letra_blanca = Font(
        color='FFFFFF',
        bold=True
    )

    for cell in ws[1]:

        cell.fill = color_encabezado

        cell.font = letra_blanca

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:

                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    wb.save(ruta_archivo)

    return send_file(
        ruta_archivo,
        as_attachment=True
    )